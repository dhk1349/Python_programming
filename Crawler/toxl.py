# -*- coding: utf-8 -*-
"""
Created on Mon Jun 29 22:13:29 2020

@author: dhk13
"""

import openpyxl


"""This section is about reading excel file"""
wb = openpyxl.load_workbook('./testfile.xlsx')
sheet = wb['Sheet1']
print(sheet.cell(row=1, column=1).value)
print("%s" % sheet.max_column)

sheet_range=sheet['A1':'B2']
for i in sheet_range:
    for j in i:
        print(j.value)
        
        
col_range=sheet['A:B']
for i in col_range:
    for j in i:
        print(j.value)
    
"""This section is about writing excel file"""
wb=openpyxl.Workbook()
sheet2=wb.create_sheet("sheet2")
sheet3=wb.create_sheet("sheet3",0)

sheet1=wb['Sheet']
sheet1['A1']="HelloWorld"

sheet2.title="두번째 시트"
wb.save('./makefile.xlsx')
wb.close()